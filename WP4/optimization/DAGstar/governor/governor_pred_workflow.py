import numpy as np
import random
import copy
import math
import json
import networkx as nx
import openpyxl

def get_value_from_excel(file_path, operator, cloud):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Find the column index for the given cloud
    cloud_column_index = None
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=2):
        for cell in col:
            if cell.value == cloud:
                cloud_column_index = cell.column
                break
    
    if cloud_column_index is None:
        raise ValueError(f"Cloud '{cloud}' not found in the sheet")
    
    # Find the row index for the given operator
    operator_row_index = None
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == operator:
                operator_row_index = cell.row
                break
    
    if operator_row_index is None:
        raise ValueError(f"Operator '{operator}' not found in the sheet")
    
    # Get the value from the appropriate cell
    value = sheet.cell(row=operator_row_index, column=cloud_column_index).value
    
    # Multiply by a factor of 100 (like the DAG approach)
    return value * 100

def get_variables(tmp_variables, tmp):
    return tmp_variables.get(tmp, "Invalid tmp value")

def read_json_file(filename):
    # Initialize empty lists to store the data
    from_list = []
    to_list = []
    latency_list = []

    # Read the JSON file
    with open(filename, 'r') as file:
        data = json.load(file)

    # Iterate over the links and extract data
    for link in data['links']:
        from_list.append(link['from'])
        to_list.append(link['to'])
        latency_list.append(link['latency'])

    return from_list, to_list, latency_list


def read_first_row(file_path):
    # Open the workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Select the active sheet
    sheet = workbook.active
    
    # Read the first row
    first_row = sheet[1]
    
    # Store values of non-empty cells in a list
    values = [cell.value for cell in first_row if cell.value is not None]
    
    return values

def dfs_shortest_path(graph, start, target):
    def dfs(node, target, visited, path, total_weight):
        nonlocal iterations  # Allows us to modify the iterations variable
        iterations += 1  # Increment the number of iterations
        visited.add(node)
        path.append(node)
        
        if node == target:
            all_paths.append((list(path), total_weight))
        else:
            for neighbor in graph.neighbors(node):
                if neighbor not in visited:
                    edge_weight = graph[node][neighbor]['weight']
                    dfs(neighbor, target, visited, path, total_weight + edge_weight)
        
        path.pop()
        visited.remove(node)
    
    all_paths = []
    visited = set()
    iterations = 0  # Initialize the iteration counter
    dfs(start, target, visited, [], 0)
    
    # Find the path with the minimum weight
    if all_paths:
        shortest_path, min_weight = min(all_paths, key=lambda x: x[1])
        return shortest_path, min_weight, iterations
    else:
        return None, float('inf'), iterations  # No path found

# Random Mapping of rangeFilter, bloomFilter, interpolation, and source to the nodes in the path
def map_operators_to_path(iot_operators, non_iot_operators, path, start_node):
    # If fewer than 4 nodes are in the path, we allow multiple operators on the same node
    if len(path) == 0:
        raise ValueError("The path cannot be empty.")
    
    # Randomly select four nodes from the path, allowing for repetition if there are fewer than 4 nodes
    selected_nodes = random.choices(path, k = len(iot_operators))  # choices allows for duplicates
    
    # Assign each operator to a node (allowing multiple operators per node if needed)
    operator_mapping = dict(zip(iot_operators, selected_nodes))

    for i in range(0, len(non_iot_operators)):
        operator_mapping[non_iot_operators[i]] = 'cloud'
    
    operator_mapping['source'] = start_node

    return operator_mapping

def compute_shortest_path_dijkstra(G, source, target):
    try:
        path = nx.dijkstra_path(G, source=source, target=target, weight='weight')
        cost = nx.dijkstra_path_length(G, source=source, target=target, weight='weight')
        return path, cost
    except nx.NetworkXNoPath:
        return None, float('inf')

def create_com_graph(node1_lst, node2_lst, com_lat_list):
    G = nx.Graph()
    for i in range(0, len(node1_lst)):
        G.add_edge(node1_lst[i], node2_lst[i], weight = com_lat_list[i])
    return G


# Define the workflow operators
operators = ["average", "blobRead", "decisionTree", "errorEstimate", "mqttPublish", "multiVarLinearReg", "senMLParse", "sink", "source"]

# Define the number of physical nodes (network nodes)
num_nodes = 7
random.seed(19)
#random.seed(517)

# Define the workflow edges (workflow)
workflow_edges = [
    ("source", "senMLParse"),
    ("source", "blobRead"),
    ("senMLParse", "decisionTree"),
    ("senMLParse", "multiVarLinearReg"),
    ("senMLParse", "average"),
    ("blobRead", "decisionTree"),
    ("blobRead", "multiVarLinearReg"),
    ("multiVarLinearReg", "errorEstimate"),
    ("average", "errorEstimate"),
    ("errorEstimate", "mqttPublish"),
    ("decisionTree", "mqttPublish"),
    ("mqttPublish", "sink")
]

iot_devices_op = ["errorEstimate", "source"]
non_iot_op = ["average", "senMLParse", "blobRead", "multiVarLinearReg", "decisionTree", "mqttPublish", "sink"]


# Map operators to indices, e.g., bloomFilter -> 0, interpolation -> 1 etc.
operator_indices = {op: idx for idx, op in enumerate(operators)}

# Convert workflow edges to indices, e.g., (5, 8), (8, 3), ...
workflow_edges_idx = [(operator_indices[op1], operator_indices[op2]) for op1, op2 in workflow_edges]

cloud_nodes = 'cloud'
if num_nodes == 7:
    no_cloud_nodes = 'cloud-0-0', 'cloud-0-1', 'cloud-1-0', 'cloud-1-1'
elif num_nodes == 15:
    no_cloud_nodes = 'cloud-0-0-0', 'cloud-0-0-1', 'cloud-0-1-0', 'cloud-0-1-1', 'cloud-1-0-0', 'cloud-1-0-1', 'cloud-1-1-0', 'cloud-1-1-1'
else:
    no_cloud_nodes = 'cloud-0-0-0-0', 'cloud-0-0-0-1', 'cloud-0-0-1-0', 'cloud-0-0-1-1', 'cloud-0-1-0-0', 'cloud-0-1-0-1', 'cloud-0-1-1-0', 'cloud-0-1-1-1', 'cloud-1-0-0-0' ,'cloud-1-0-0-1', 'cloud-1-0-1-0', 'cloud-1-0-1-1', 'cloud-1-1-0-0', 'cloud-1-1-0-1', 'cloud-1-1-1-0', 'cloud-1-1-1-1'


# Communication costs part
# Extract the information regarding communications latencies for this simulation from the appropriate JSON file
json_file_path = "../network_sample/pred_dataflow/" + str(num_nodes) + "_1/links_" + str(num_nodes) + "_1.json"
from_list, to_list, latency_list = read_json_file(json_file_path)

# Create the communication graph
global COM_GRAPH
COM_GRAPH = create_com_graph(from_list, to_list, latency_list)

# The file that contains the data of latencies of executing some operator at some network node
file_path = "../datasets/pred_" + str(num_nodes) + "_dataflow.xlsx"

# Compute shortest path using DFS between two cloud nodes
path_nodes = []
total_it = 0
for i in range(0, 1):
    start_node = random.choice(no_cloud_nodes)
    print(start_node)
    if num_nodes == 31 or num_nodes == 15 or num_nodes == 7:
        start_node = 'cloud-0-0'    
    
    target_node = "cloud"
    path, cost, it = dfs_shortest_path(COM_GRAPH, start_node, target_node)
    total_it += it
    path_nodes = list(set(path_nodes + path))
#path_nodes.remove('cloud')
path_nodes.sort()
print(f"Nodes are {path_nodes}")

# Randomly map rangeFilter, bloomFilter, interpolation, and source to nodes in the path
operator_mapping = map_operators_to_path(iot_devices_op, non_iot_op, path_nodes, start_node)
print(f"Operator Mapping: {operator_mapping}")

total_latency = 0

###########################################
# SUM scenario
###########################################

# total_latency += get_value_from_excel(file_path, 'source', operator_mapping['source'])

# tmp1 = total_latency
# tmp2 = total_latency


# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['blobRead'])
# tmp1 += tmp
# tmp1 += get_value_from_excel(file_path, 'blobRead', operator_mapping['blobRead'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['blobRead'], operator_mapping['decisionTree'])
# tmp1 += tmp

# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['senMLParse'])
# tmp2 += tmp
# tmp2 += get_value_from_excel(file_path, 'senMLParse', operator_mapping['senMLParse'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['senMLParse'], operator_mapping['decisionTree'])
# tmp2 += tmp

# tmp3 = tmp1 + tmp2
# tmp1 = total_latency
# tmp2 = total_latency

# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['blobRead'])
# tmp1 += tmp
# tmp1 += get_value_from_excel(file_path, 'blobRead', operator_mapping['blobRead'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['blobRead'], operator_mapping['multiVarLinearReg'])
# tmp1 += tmp

# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['senMLParse'])
# tmp2 += tmp
# tmp2 += get_value_from_excel(file_path, 'senMLParse', operator_mapping['senMLParse'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['senMLParse'], operator_mapping['multiVarLinearReg'])
# tmp2 += tmp

# tmp4 = tmp1 + tmp2
# tmp2 = total_latency

# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['senMLParse'])
# tmp2 += tmp
# tmp2 += get_value_from_excel(file_path, 'senMLParse', operator_mapping['senMLParse'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['senMLParse'], operator_mapping['average'])
# tmp2 += tmp

# tmp2 += get_value_from_excel(file_path, 'average', operator_mapping['average'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['average'], operator_mapping['errorEstimate'])
# tmp2 += tmp

# tmp4 += get_value_from_excel(file_path, 'multiVarLinearReg', operator_mapping['senMLParse'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['senMLParse'], operator_mapping['errorEstimate'])
# tmp4 += tmp

# tmp5 = tmp2 + tmp4
# tmp5 += get_value_from_excel(file_path, 'errorEstimate', operator_mapping['errorEstimate'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['errorEstimate'], operator_mapping['mqttPublish'])
# tmp5 += tmp

# tmp3 += get_value_from_excel(file_path, 'decisionTree', operator_mapping['decisionTree'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['decisionTree'], operator_mapping['mqttPublish'])
# tmp3 += tmp

# tmp6 = tmp3 + tmp5
# total_latency = tmp6

# total_latency += get_value_from_excel(file_path, 'mqttPublish', operator_mapping['mqttPublish'])
# _, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['mqttPublish'], operator_mapping['sink'])
# total_latency += tmp
# total_latency += get_value_from_excel(file_path, 'sink', operator_mapping['sink'])


###########################################
# MAX scenario
###########################################

total_latency += get_value_from_excel(file_path, 'source', operator_mapping['source'])

tmp1 = total_latency
tmp2 = total_latency

_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['blobRead'])
tmp1 += tmp
tmp1 += get_value_from_excel(file_path, 'blobRead', operator_mapping['blobRead'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['blobRead'], operator_mapping['decisionTree'])
tmp1 += tmp

_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['senMLParse'])
tmp2 += tmp
tmp2 += get_value_from_excel(file_path, 'senMLParse', operator_mapping['senMLParse'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['senMLParse'], operator_mapping['decisionTree'])
tmp2 += tmp

tmp3 = max(tmp1, tmp2)
tmp1 = total_latency
tmp2 = total_latency

_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['blobRead'])
tmp1 += tmp
tmp1 += get_value_from_excel(file_path, 'blobRead', operator_mapping['blobRead'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['blobRead'], operator_mapping['multiVarLinearReg'])
tmp1 += tmp

_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['senMLParse'])
tmp2 += tmp
tmp2 += get_value_from_excel(file_path, 'senMLParse', operator_mapping['senMLParse'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['senMLParse'], operator_mapping['multiVarLinearReg'])
tmp2 += tmp

tmp4 = max(tmp1, tmp2)
tmp2 = total_latency

_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['source'], operator_mapping['senMLParse'])
tmp2 += tmp
tmp2 += get_value_from_excel(file_path, 'senMLParse', operator_mapping['senMLParse'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['senMLParse'], operator_mapping['average'])
tmp2 += tmp

tmp2 += get_value_from_excel(file_path, 'average', operator_mapping['average'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['average'], operator_mapping['errorEstimate'])
tmp2 += tmp

tmp4 += get_value_from_excel(file_path, 'multiVarLinearReg', operator_mapping['multiVarLinearReg'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['multiVarLinearReg'], operator_mapping['errorEstimate'])
tmp4 += tmp

tmp5 = max(tmp2, tmp4)
tmp5 += get_value_from_excel(file_path, 'errorEstimate', operator_mapping['errorEstimate'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['errorEstimate'], operator_mapping['mqttPublish'])
tmp5 += tmp

tmp3 += get_value_from_excel(file_path, 'decisionTree', operator_mapping['decisionTree'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['decisionTree'], operator_mapping['mqttPublish'])
tmp3 += tmp

tmp6 = max(tmp3, tmp5)
total_latency = tmp6

total_latency += get_value_from_excel(file_path, 'mqttPublish', operator_mapping['mqttPublish'])
_, tmp = compute_shortest_path_dijkstra(COM_GRAPH, operator_mapping['mqttPublish'], operator_mapping['sink'])
total_latency += tmp
total_latency += get_value_from_excel(file_path, 'sink', operator_mapping['sink'])


# Print the latency of the resulted topology
print(f"\nThe latency of this topology is: {total_latency}")
print(f"\nIteration for Governor approach is: {total_it}")
